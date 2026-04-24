from __future__ import annotations

import ast
import json
import math
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Optional
import uuid

from sqlmodel import Session, select

from app.models.contact import Contact, ContactType
from app.models.expense import Expense, ExpenseCategory
from app.models.mileage import MileageLog
from app.models.order import Order, OrderItem, OrderStatus, PaymentStatus
from app.models.user import User


EXCEL_EPOCH = datetime(1899, 12, 30, tzinfo=timezone.utc)


@dataclass
class ImportCounts:
    contacts_created: int = 0
    contacts_matched: int = 0
    orders_created: int = 0
    expenses_created: int = 0
    mileage_created: int = 0
    orders_skipped_as_quotes: int = 0
    skipped_empty_rows: int = 0


@dataclass
class ImportWarnings:
    items: list[str] = field(default_factory=list)

    def add(self, message: str) -> None:
        if message not in self.items:
            self.items.append(message)


@dataclass
class ImportResult:
    counts: ImportCounts
    warnings: ImportWarnings


class MarvelousCreationsImporter:
    def __init__(self, session: Session, current_user: User):
        self.session = session
        self.current_user = current_user
        self.counts = ImportCounts()
        self.warnings = ImportWarnings()
        self._contact_cache: dict[str, Contact] = {}

    def import_workbook(self, workbook_path: str | Path) -> ImportResult:
        sheets = load_workbook_rows(workbook_path)
        self.import_sheets(
            contacts_rows=sheets.get("Contacts", []),
            orders_rows=sheets.get("Orders", []),
            expenses_rows=sheets.get("Expenses", []),
            mileage_rows=sheets.get("Mileage", []),
        )
        return ImportResult(counts=self.counts, warnings=self.warnings)

    def import_sheets(
        self,
        *,
        contacts_rows: Iterable[dict[str, Any]],
        orders_rows: Iterable[dict[str, Any]],
        expenses_rows: Iterable[dict[str, Any]],
        mileage_rows: Iterable[dict[str, Any]],
    ) -> ImportResult:
        for row in contacts_rows:
            if row_is_empty(row):
                self.counts.skipped_empty_rows += 1
                continue
            self._resolve_contact(row)

        for row in orders_rows:
            if row_is_empty(row):
                self.counts.skipped_empty_rows += 1
                continue
            self._import_order(row)

        for row in expenses_rows:
            if row_is_empty(row):
                self.counts.skipped_empty_rows += 1
                continue
            self._import_expense(row)

        for row in mileage_rows:
            if row_is_empty(row):
                self.counts.skipped_empty_rows += 1
                continue
            self._import_mileage(row)

        self.session.commit()
        return ImportResult(counts=self.counts, warnings=self.warnings)

    def _resolve_contact(self, row: dict[str, Any]) -> Optional[Contact]:
        email = normalize_email(first_present(row, "EmailAddress", "ContactEmail", "Email"))
        phone = normalize_phone(first_present(row, "Number", "Phone", "ContactPhone"))
        contact_name = cleaned_string(first_present(row, "Contact", "FullName", "Name"))
        first_name = cleaned_string(first_present(row, "FirstName"))
        last_name = cleaned_string(first_present(row, "LastName"))
        company_name = cleaned_string(first_present(row, "ContactCompany", "Company", "CompanyName"))

        if not first_name and not last_name and contact_name:
            first_name, last_name = split_name(contact_name)

        cache_keys = [key for key in [email, phone, normalized_name_key(first_name, last_name, company_name)] if key]
        for key in cache_keys:
            cached = self._contact_cache.get(key)
            if cached is not None:
                self.counts.contacts_matched += 1
                return cached

        statement = select(Contact).where(Contact.user_id == self.current_user.id)
        candidates = self.session.exec(statement).all()
        for candidate in candidates:
            if email and normalize_email(candidate.email) == email:
                self._cache_contact(candidate)
                self.counts.contacts_matched += 1
                return candidate
            if phone and normalize_phone(candidate.phone) == phone:
                self._cache_contact(candidate)
                self.counts.contacts_matched += 1
                return candidate
            if normalized_name_key(candidate.first_name, candidate.last_name, candidate.company_name) == normalized_name_key(first_name, last_name, company_name):
                self._cache_contact(candidate)
                self.counts.contacts_matched += 1
                return candidate

        address = parse_address(first_present(row, "Address", "StreetAddress"))
        notes = build_contact_notes(row)
        contact = Contact(
            user_id=self.current_user.id,
            first_name=first_name,
            last_name=last_name,
            company_name=company_name,
            email=email,
            phone=phone,
            address_line1=address.get("address_line1"),
            address_line2=address.get("address_line2"),
            city=address.get("city"),
            state_province=address.get("state_province"),
            postal_code=address.get("postal_code"),
            country=address.get("country") or "US",
            contact_type=ContactType.CUSTOMER,
            notes=notes,
        )
        self.session.add(contact)
        self.session.flush()
        self._cache_contact(contact)
        self.counts.contacts_created += 1
        return contact

    def _cache_contact(self, contact: Contact) -> None:
        for key in [
            normalize_email(contact.email),
            normalize_phone(contact.phone),
            normalized_name_key(contact.first_name, contact.last_name, contact.company_name),
        ]:
            if key:
                self._contact_cache[key] = contact

    def _import_order(self, row: dict[str, Any]) -> None:
        is_quote = coerce_bool(first_present(row, "IsQuote"))
        if is_quote:
            self.counts.orders_skipped_as_quotes += 1
            return

        order_number = cleaned_string(first_present(row, "OrderNumber"))
        if not order_number:
            self.warnings.add("Skipped order row without OrderNumber.")
            return

        existing = self.session.exec(
            select(Order).where(
                Order.user_id == self.current_user.id,
                Order.order_number == order_number,
            )
        ).first()
        if existing is not None:
            self.warnings.add(f"Skipped duplicate legacy order_number {order_number}.")
            return

        contact = self._resolve_contact(row)
        order_dt = coerce_datetime(first_present(row, "OrderDate")) or datetime.now(timezone.utc)
        due_dt = coerce_datetime(
            first_present(row, "DueDate", "PickupDate", "DeliveryDate", "EventDate")
        ) or order_dt

        subtotal = coerce_money(
            first_present(row, "Subtotal", "SubTotal", "SubTotalAmount", "ProductsTotal")
        ) or 0.0
        setup_delivery_amount = coerce_money(first_present(row, "SetupDeliveryAmount")) or 0.0
        tax = aggregate_tax(row)
        total_amount = coerce_money(first_present(row, "Total", "TotalAmount", "GrandTotal"))
        if total_amount is None:
            total_amount = round(subtotal + setup_delivery_amount + tax, 2)
        deposit_amount = coerce_money(first_present(row, "DepositAmount", "Deposit"))
        explicit_amount_paid = coerce_money(first_present(row, "AmountPaid", "PaidAmount", "PaymentsReceived"))
        amount_paid = explicit_amount_paid or 0.0
        explicit_balance_due = coerce_money(first_present(row, "BalanceDue", "RemainingBalance"))
        balance_due = explicit_balance_due
        if balance_due is None:
            balance_due = round(total_amount - amount_paid, 2)

        payment_status = infer_payment_status(
            row=row,
            total_amount=total_amount,
            deposit_amount=deposit_amount,
            amount_paid=amount_paid,
            balance_due=balance_due,
            has_explicit_amount_paid=explicit_amount_paid is not None,
            has_explicit_balance_due=explicit_balance_due is not None,
        )
        status, payment_status = infer_statuses(
            raw_status=first_present(row, "OrderStatusId"),
            due_dt=due_dt,
            total_amount=total_amount,
            balance_due=balance_due,
            amount_paid=amount_paid,
            payment_status=payment_status,
        )

        internal_notes = build_order_internal_notes(
            row,
            status_raw=first_present(row, "OrderStatusId"),
            mapped_status=status,
        )
        notes_to_customer = cleaned_string(first_present(row, "NotesToCustomer"))
        delivery_method = cleaned_string(first_present(row, "DeliveryMethod"))
        if not delivery_method:
            delivery_method = infer_delivery_method(row, setup_delivery_amount=setup_delivery_amount)

        order = Order(
            user_id=self.current_user.id,
            customer_contact_id=contact.id if contact else None,
            customer_name=compose_contact_name(contact, row),
            customer_email=normalize_email(first_present(row, "ContactEmail", "CustomerEmail")) or (str(contact.email) if contact and contact.email else None),
            customer_phone=normalize_phone(first_present(row, "ContactPhone", "Number", "Phone")) or (contact.phone if contact else None),
            order_number=order_number,
            status=status,
            payment_status=payment_status,
            order_date=order_dt,
            due_date=due_dt,
            delivery_method=delivery_method,
            subtotal=subtotal,
            tax=tax,
            total_amount=total_amount,
            deposit_amount=deposit_amount,
            balance_due=balance_due,
            deposit_due_date=coerce_date(first_present(row, "DepositDueDate")),
            balance_due_date=coerce_date(first_present(row, "BalanceDueDate")),
            notes_to_customer=notes_to_customer,
            internal_notes=internal_notes,
        )
        self.session.add(order)
        self.session.flush()

        items = parse_order_items(row, subtotal=subtotal, total_amount=total_amount)
        for item in items:
            self.session.add(
                OrderItem(
                    user_id=self.current_user.id,
                    order_id=order.id,
                    name=item["name"],
                    description=item.get("description"),
                    quantity=item["quantity"],
                    unit_price=item["unit_price"],
                    total_price=item["total_price"],
                )
            )

        self.counts.orders_created += 1

    def _import_expense(self, row: dict[str, Any]) -> None:
        description = cleaned_string(first_present(row, "Expense", "Description", "Name"))
        amount = coerce_money(first_present(row, "Amount", "Cost", "Total"))
        expense_date = coerce_date(first_present(row, "ExpenseDate", "Date", "TransactionDate"))
        if not description or amount is None or expense_date is None:
            self.warnings.add("Skipped expense row missing description, amount, or date.")
            return

        expense = Expense(
            user_id=self.current_user.id,
            date=expense_date,
            description=description,
            amount=amount,
            category=map_expense_category(first_present(row, "Category", "ExpenseCategory")),
            vendor=cleaned_string(first_present(row, "Vendor", "Payee", "Store")),
            notes=join_note_parts(
                cleaned_string(first_present(row, "Notes")),
                metadata_lines(row, ["ExpenseID", "ReceiptNumber"]),
            ),
        )
        self.session.add(expense)
        self.counts.expenses_created += 1

    def _import_mileage(self, row: dict[str, Any]) -> None:
        mileage_date = coerce_date(first_present(row, "MileageDate", "Date", "TripDate"))
        distance = coerce_float(first_present(row, "Distance", "Miles", "Mileage"))
        if mileage_date is None or distance is None:
            self.warnings.add("Skipped mileage row missing date or distance.")
            return

        reimbursement_rate = coerce_float(first_present(row, "Rate", "ReimbursementRate"))
        reimbursement_amount = None
        if reimbursement_rate is not None:
            reimbursement_amount = round(distance * reimbursement_rate, 2)

        log = MileageLog(
            user_id=self.current_user.id,
            date=mileage_date,
            start_location=cleaned_string(first_present(row, "StartLocation", "From")),
            end_location=cleaned_string(first_present(row, "EndLocation", "To")),
            distance=distance,
            purpose=cleaned_string(first_present(row, "Purpose", "Reason")),
            vehicle_identifier=cleaned_string(first_present(row, "Vehicle", "VehicleIdentifier")),
            notes=join_note_parts(
                cleaned_string(first_present(row, "Notes")),
                metadata_lines(row, ["MileageID"]),
            ),
            reimbursement_rate=reimbursement_rate,
            reimbursement_amount=reimbursement_amount,
        )
        self.session.add(log)
        self.counts.mileage_created += 1


def load_workbook_rows(workbook_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read Marvelous Creations XLSX files.") from exc

    workbook = load_workbook(filename=str(workbook_path), data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            sheets[sheet_name] = []
            continue
        headers = [cleaned_string(cell) or "" for cell in rows[0]]
        sheet_rows: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            row = {headers[index]: value for index, value in enumerate(raw_row) if index < len(headers) and headers[index]}
            sheet_rows.append(row)
        sheets[sheet_name] = sheet_rows
    return sheets


def row_is_empty(row: dict[str, Any]) -> bool:
    return all(cleaned_string(value) is None for value in row.values())


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and cleaned_string(row[key]) is not None:
            return row[key]
    return None


def cleaned_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.upper() == "NULL":
            return None
        return text
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    if not text or text.upper() == "NULL":
        return None
    return text


def normalize_legacy_numeric_status(value: Any) -> Optional[str]:
    text = cleaned_string(value)
    if text is None:
        return None
    try:
        numeric = float(text)
    except ValueError:
        return text.lower()
    if math.isnan(numeric):
        return None
    if numeric.is_integer():
        return str(int(numeric))
    return text.lower()



def coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return False
        return value != 0
    text = cleaned_string(value)
    if text is None:
        return False
    return text.lower() in {"1", "1.0", "true", "yes", "y"}


def normalize_email(value: Any) -> Optional[str]:
    text = cleaned_string(value)
    return text.lower() if text else None


def normalize_phone(value: Any) -> Optional[str]:
    text = cleaned_string(value)
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    return text


def coerce_money(value: Any) -> Optional[float]:
    text = cleaned_string(value)
    if text is None:
        return None
    normalized = text.replace("$", "").replace(",", "").strip()
    try:
        return round(float(Decimal(normalized)), 2)
    except (InvalidOperation, ValueError):
        return None


def coerce_float(value: Any) -> Optional[float]:
    text = cleaned_string(value)
    if text is None:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def coerce_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        whole_days = float(value)
        return EXCEL_EPOCH + timedelta(days=whole_days)
    text = cleaned_string(value)
    if text is None:
        return None
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M %p",
    ]:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(timezone.utc)
    except ValueError:
        return None


def coerce_date(value: Any) -> Optional[date]:
    dt = coerce_datetime(value)
    return dt.date() if dt else None


def aggregate_tax(row: dict[str, Any]) -> float:
    tax_fields = [
        "Tax",
        "SalesTax",
        "TaxAmount",
        "ShippingTaxAmount",
        "StateTax",
        "LocalTax",
        "TaxAmount1",
        "TaxAmount2",
        "TaxAmount3",
        "TaxAmount4",
        "TaxAmount5",
    ]
    amounts = [coerce_money(row.get(field)) for field in tax_fields if field in row]
    present = [amount for amount in amounts if amount is not None]
    if present:
        return round(sum(present), 2)
    return 0.0


def split_name(value: str | None) -> tuple[Optional[str], Optional[str]]:
    text = cleaned_string(value)
    if not text:
        return None, None
    pieces = text.split(maxsplit=1)
    if len(pieces) == 1:
        return pieces[0], None
    return pieces[0], pieces[1]


def normalized_name_key(first_name: Optional[str], last_name: Optional[str], company_name: Optional[str]) -> Optional[str]:
    parts = [cleaned_string(part) for part in [first_name, last_name, company_name]]
    joined = "|".join(part.lower() for part in parts if part)
    return joined or None


def compose_contact_name(contact: Optional[Contact], row: dict[str, Any]) -> Optional[str]:
    if contact:
        parts = [part for part in [contact.first_name, contact.last_name] if part]
        if parts:
            return " ".join(parts)
        if contact.company_name:
            return contact.company_name
    return cleaned_string(first_present(row, "Contact", "CustomerName"))


def parse_address(value: Any) -> dict[str, Optional[str]]:
    text = cleaned_string(value)
    if not text:
        return {}
    lines = [part.strip() for part in text.replace("\r", "\n").split("\n") if part.strip()]
    address_line1 = lines[0] if lines else text
    address_line2 = lines[1] if len(lines) > 2 else None
    city = state_province = postal_code = None
    tail = lines[-1] if len(lines) > 1 else None
    if tail and "," in tail:
        city_part, state_zip = [chunk.strip() for chunk in tail.split(",", 1)]
        city = city_part or None
        state_bits = state_zip.split()
        if state_bits:
            state_province = state_bits[0]
        if len(state_bits) > 1:
            postal_code = " ".join(state_bits[1:])
    return {
        "address_line1": address_line1,
        "address_line2": address_line2,
        "city": city,
        "state_province": state_province,
        "postal_code": postal_code,
        "country": "US",
    }


def build_contact_notes(row: dict[str, Any]) -> Optional[str]:
    return join_note_parts(
        cleaned_string(first_present(row, "Notes")),
        metadata_lines(row, ["ContactID"]),
    )


def build_order_internal_notes(row: dict[str, Any], *, status_raw: Any, mapped_status: OrderStatus | None = None) -> Optional[str]:
    extra: dict[str, Any] = {"legacy_status_raw": cleaned_string(status_raw)}
    if mapped_status is not None:
        extra["bakemate_status"] = mapped_status.value
    return join_note_parts(
        cleaned_string(first_present(row, "Notes")),
        cleaned_string(first_present(row, "JobSheetNotes")),
        cleaned_string(first_present(row, "ThemeDetails")),
        cleaned_string(first_present(row, "EventType")),
        metadata_lines(
            row,
            [
                "OrderStatusId",
                "Year",
                "DepositAmount",
                "Deposit",
                "AmountPaid",
                "PaidAmount",
                "PaymentsReceived",
                "BalanceDue",
                "RemainingBalance",
                "DepositDueDate",
                "BalanceDueDate",
            ],
            extra=extra,
        ),
    )


def metadata_lines(
    row: dict[str, Any],
    keys: list[str],
    extra: Optional[dict[str, Any]] = None,
) -> Optional[str]:
    lines: list[str] = []
    for key in keys:
        value = cleaned_string(row.get(key))
        if value is not None:
            lines.append(f"Legacy {key}: {value}")
    if extra:
        for key, value in extra.items():
            text = cleaned_string(value)
            if text is not None:
                lines.append(f"Legacy {key}: {text}")
    return "\n".join(lines) if lines else None


def join_note_parts(*parts: Optional[str]) -> Optional[str]:
    cleaned = [part.strip() for part in parts if part and part.strip()]
    return "\n\n".join(cleaned) if cleaned else None


def infer_delivery_method(row: dict[str, Any], *, setup_delivery_amount: float = 0.0) -> Optional[str]:
    text = cleaned_string(first_present(row, "Delivery", "PickupOrDelivery"))
    if text:
        lowered = text.lower()
        if "deliver" in lowered:
            return "delivery"
        if "pick" in lowered:
            return "pickup"
        return text

    if setup_delivery_amount > 0:
        return "delivery"

    delivery_date = coerce_datetime(first_present(row, "DeliveryDate", "DeliveryTime"))
    pickup_date = coerce_datetime(first_present(row, "PickupDate", "PickupTime"))
    if delivery_date and not pickup_date:
        return "delivery"
    if pickup_date and not delivery_date:
        return "pickup"

    address_text = " ".join(
        filter(
            None,
            [
                cleaned_string(first_present(row, "DeliveryAddress", "Address", "StreetAddress")),
                cleaned_string(first_present(row, "DeliveryCity", "City")),
            ],
        )
    ).lower()
    if address_text and any(token in address_text for token in ["deliver", "delivery", "drop off"]):
        return "delivery"

    clue_fields = [
        "Notes",
        "JobSheetNotes",
        "ThemeDetails",
        "EventType",
        "SetupDeliveryNotes",
        "DeliveryInstructions",
    ]
    clues = " ".join(cleaned_string(row.get(field)) or "" for field in clue_fields).lower()
    if any(token in clues for token in ["porch pickup", "pick up", "pickup"]):
        return "pickup"
    if any(token in clues for token in ["deliver", "delivery", "drop off", "setup on site"]):
        return "delivery"
    return None


def map_expense_category(value: Any) -> ExpenseCategory:
    text = (cleaned_string(value) or "").lower()
    if any(token in text for token in ["ingredient", "flour", "sugar", "food"]):
        return ExpenseCategory.INGREDIENTS
    if any(token in text for token in ["supply", "packaging", "box"]):
        return ExpenseCategory.SUPPLIES
    if any(token in text for token in ["utility", "electric", "gas", "water"]):
        return ExpenseCategory.UTILITIES
    if "rent" in text:
        return ExpenseCategory.RENT
    if any(token in text for token in ["market", "ad", "promo"]):
        return ExpenseCategory.MARKETING
    if any(token in text for token in ["fee", "stripe", "square", "processing"]):
        return ExpenseCategory.FEES
    return ExpenseCategory.OTHER


# Staging validation against the real workbook showed that low numeric values like
# `2.0` appear on many non-quote orders. Since quote rows are already excluded via
# `IsQuote`, treat ambiguous low numeric statuses conservatively and keep them in a
# believable active-order bucket instead of resurfacing them as `quote_sent`.
LEGACY_NUMERIC_ORDER_STATUSES: dict[str, OrderStatus] = {
    "1": OrderStatus.INQUIRY,
    "2": OrderStatus.CONFIRMED,
    "3": OrderStatus.CONFIRMED,
    "4": OrderStatus.IN_PROGRESS,
    "5": OrderStatus.READY_FOR_PICKUP,
    "6": OrderStatus.COMPLETED,
    "7": OrderStatus.CONFIRMED,
    "8": OrderStatus.CANCELLED,
}

# Real staging reruns showed old legacy workbook rows with `0.0` / `2.0` status
# values clustering in a vague "active-ish" bucket. A first-pass direct mapping to
# `confirmed` preserved safety for live/future jobs, but it flattened too many
# clearly historical rows into the same state. Keep the refinement intentionally
# conservative: only revisit past-due ambiguous rows, and only when payment/age
# gives us a believable stronger signal.
LEGACY_HISTORICAL_AMBIGUOUS_STATUSES = {"0", "2"}
LEGACY_HISTORICAL_STATUS_AGE_DAYS = 30
LEGACY_HISTORICAL_ZERO_BALANCE_AGE_DAYS = 180
LEGACY_HISTORICAL_ANCIENT_AMBIGUOUS_AGE_DAYS = 730

GENERIC_ITEM_NAMES = {
    "item",
    "items",
    "other",
    "product",
    "products",
    "recipe",
    "custom",
    "misc",
    "miscellaneous",
}


def infer_payment_status(
    *,
    row: dict[str, Any],
    total_amount: float,
    deposit_amount: float | None,
    amount_paid: float,
    balance_due: float,
    has_explicit_amount_paid: bool,
    has_explicit_balance_due: bool,
) -> PaymentStatus:
    if balance_due <= 0 and total_amount > 0:
        return PaymentStatus.PAID_IN_FULL

    if amount_paid > 0:
        return PaymentStatus.DEPOSIT_PAID

    if has_explicit_balance_due and total_amount > 0 and balance_due < total_amount:
        return PaymentStatus.DEPOSIT_PAID

    if (
        has_explicit_balance_due
        and deposit_amount is not None
        and deposit_amount > 0
        and total_amount > 0
        and abs(balance_due - round(total_amount - deposit_amount, 2)) < 0.01
    ):
        return PaymentStatus.DEPOSIT_PAID

    note_fields = [
        "Notes",
        "JobSheetNotes",
        "NotesToCustomer",
        "ThemeDetails",
        "PaymentNotes",
        "PaymentStatus",
    ]
    payment_text = " ".join(cleaned_string(row.get(field)) or "" for field in note_fields).lower()
    if any(token in payment_text for token in ["paid in full", "pif", "payment complete", "fully paid"]):
        return PaymentStatus.PAID_IN_FULL
    if any(token in payment_text for token in ["deposit paid", "retainer paid", "partial payment received"]):
        return PaymentStatus.DEPOSIT_PAID

    return PaymentStatus.UNPAID



def infer_statuses(
    *,
    raw_status: Any,
    due_dt: datetime,
    total_amount: float,
    balance_due: float,
    amount_paid: float,
    payment_status: PaymentStatus,
) -> tuple[OrderStatus, PaymentStatus]:
    text = cleaned_string(raw_status)
    now = datetime.now(timezone.utc)

    normalized_numeric_status = normalize_legacy_numeric_status(text)
    if normalized_numeric_status in {"5", "6", "7"} and due_dt < now:
        if payment_status == PaymentStatus.PAID_IN_FULL:
            return OrderStatus.COMPLETED, payment_status
        if normalized_numeric_status == "5":
            return OrderStatus.READY_FOR_PICKUP, payment_status
        if payment_status == PaymentStatus.UNPAID:
            return OrderStatus.IN_PROGRESS, payment_status

    if text:
        lowered = text.lower()
        explicit = {
            "cancel": OrderStatus.CANCELLED,
            "complete": OrderStatus.COMPLETED,
            "pickup": OrderStatus.READY_FOR_PICKUP,
            "progress": OrderStatus.IN_PROGRESS,
            "quote": OrderStatus.QUOTE_SENT,
            "inquiry": OrderStatus.INQUIRY,
            "confirm": OrderStatus.CONFIRMED,
        }
        for token, status in explicit.items():
            if token in lowered:
                return status, payment_status

        numeric_status = LEGACY_NUMERIC_ORDER_STATUSES.get(normalized_numeric_status or lowered)
        if numeric_status is not None:
            status = numeric_status
            historical_status = infer_historical_status_for_ambiguous_legacy_row(
                normalized_numeric_status=normalized_numeric_status,
                due_dt=due_dt,
                total_amount=total_amount,
                balance_due=balance_due,
                payment_status=payment_status,
                now=now,
            )
            if historical_status is not None:
                status = historical_status
            return status, payment_status

    historical_status = infer_historical_status_for_ambiguous_legacy_row(
        normalized_numeric_status=normalized_numeric_status,
        due_dt=due_dt,
        total_amount=total_amount,
        balance_due=balance_due,
        payment_status=payment_status,
        now=now,
    )
    if historical_status is not None:
        return historical_status, payment_status

    if due_dt < now and payment_status == PaymentStatus.PAID_IN_FULL:
        return OrderStatus.COMPLETED, payment_status
    return OrderStatus.CONFIRMED, payment_status


def infer_historical_status_for_ambiguous_legacy_row(
    *,
    normalized_numeric_status: Optional[str],
    due_dt: datetime,
    total_amount: float,
    balance_due: float,
    payment_status: PaymentStatus,
    now: datetime,
) -> Optional[OrderStatus]:
    if normalized_numeric_status not in LEGACY_HISTORICAL_AMBIGUOUS_STATUSES:
        return None
    if due_dt >= now:
        return None

    if payment_status == PaymentStatus.PAID_IN_FULL:
        return OrderStatus.COMPLETED

    age_days = (now - due_dt).days
    if age_days >= LEGACY_HISTORICAL_ZERO_BALANCE_AGE_DAYS and (
        balance_due <= 0 or total_amount <= 0
    ):
        return OrderStatus.COMPLETED

    if (
        payment_status == PaymentStatus.UNPAID
        and age_days >= LEGACY_HISTORICAL_ANCIENT_AMBIGUOUS_AGE_DAYS
    ):
        return OrderStatus.COMPLETED

    if payment_status == PaymentStatus.DEPOSIT_PAID and age_days >= LEGACY_HISTORICAL_STATUS_AGE_DAYS:
        return OrderStatus.IN_PROGRESS

    return None


def parse_order_items(row: dict[str, Any], *, subtotal: float, total_amount: float) -> list[dict[str, Any]]:
    product_items = parse_jsonish(first_present(row, "ProductItems"))
    product_recipes = parse_jsonish(first_present(row, "ProductRecipes"))
    normalized_items: list[dict[str, Any]] = []

    for source, kind in [(product_items, "item"), (product_recipes, "recipe")]:
        if isinstance(source, list):
            for entry in source:
                normalized = normalize_item_entry(entry, default_kind=kind)
                if normalized:
                    normalized_items.append(normalized)

    if not normalized_items:
        fallback_name = cleaned_string(first_present(row, "EventType")) or "Imported order"
        return [
            {
                "name": fallback_name,
                "description": cleaned_string(first_present(row, "ThemeDetails")),
                "quantity": 1,
                "unit_price": round(total_amount or subtotal, 2),
                "total_price": round(total_amount or subtotal, 2),
            }
        ]

    total_known = sum(item["total_price"] for item in normalized_items if item["total_price"] > 0)
    target_total = subtotal if subtotal > 0 else total_amount
    if target_total > 0 and total_known <= 0:
        per_item = round(target_total / len(normalized_items), 2)
        for item in normalized_items:
            item["unit_price"] = per_item / item["quantity"]
            item["total_price"] = per_item
    elif target_total > 0 and abs(total_known - target_total) >= 0.01:
        scale = target_total / total_known if total_known else 1.0
        running_total = 0.0
        for index, item in enumerate(normalized_items):
            if index == len(normalized_items) - 1:
                item_total = round(target_total - running_total, 2)
            else:
                item_total = round(item["total_price"] * scale, 2)
                running_total += item_total
            item["total_price"] = item_total
            item["unit_price"] = round(item_total / item["quantity"], 2)
    return normalized_items


def parse_jsonish(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (list, dict)):
        return value
    text = cleaned_string(value)
    if text is None:
        return None
    for loader in (json.loads, ast.literal_eval):
        try:
            return loader(text)
        except Exception:
            continue
    if "|" in text:
        return [{"name": part.strip(), "quantity": 1} for part in text.split("|") if part.strip()]
    if "," in text:
        return [{"name": part.strip(), "quantity": 1} for part in text.split(",") if part.strip()]
    return [{"name": text, "quantity": 1}]


def looks_generic_item_name(value: Optional[str], *, default_kind: str) -> bool:
    text = (cleaned_string(value) or "").strip().lower()
    if not text:
        return True
    return text in GENERIC_ITEM_NAMES or text == default_kind.lower() or text == default_kind.title().lower()



def choose_best_item_name(entry: dict[str, Any], *, default_kind: str, description: Optional[str]) -> str:
    primary_candidates = [
        entry.get("name"),
        entry.get("Name"),
        entry.get("productName"),
        entry.get("ProductName"),
        entry.get("recipeName"),
        entry.get("RecipeName"),
        entry.get("ItemName"),
        entry.get("ProductItemName"),
        entry.get("title"),
        entry.get("ProductType"),
    ]
    for candidate in primary_candidates:
        text = cleaned_string(candidate)
        if text and not looks_generic_item_name(text, default_kind=default_kind):
            return text

    secondary_candidates = [
        entry.get("Flavor"),
        entry.get("CakeFlavor"),
        entry.get("Filling"),
        entry.get("Size"),
        entry.get("TreatType"),
        entry.get("Category"),
        entry.get("ProductCategory"),
        description,
    ]
    for candidate in secondary_candidates:
        text = cleaned_string(candidate)
        if text and not looks_generic_item_name(text, default_kind=default_kind):
            return text

    for candidate in primary_candidates:
        text = cleaned_string(candidate)
        if text:
            return text
    return default_kind.title()



def normalize_item_entry(entry: Any, *, default_kind: str) -> Optional[dict[str, Any]]:
    if isinstance(entry, str):
        return {
            "name": entry.strip(),
            "description": None,
            "quantity": 1,
            "unit_price": 0.0,
            "total_price": 0.0,
        }
    if not isinstance(entry, dict):
        return None
    description = cleaned_string(
        entry.get("description") or entry.get("Description") or entry.get("Details") or entry.get("ProductDescription")
    )
    name = choose_best_item_name(entry, default_kind=default_kind, description=description)
    quantity = int(
        coerce_float(entry.get("quantity") or entry.get("Quantity") or entry.get("Qty") or 1) or 1
    )
    unit_price = coerce_money(
        entry.get("unit_price")
        or entry.get("unitPrice")
        or entry.get("price")
        or entry.get("Price")
        or entry.get("SellingPrice")
    ) or 0.0
    total_price = coerce_money(
        entry.get("total_price")
        or entry.get("totalPrice")
        or entry.get("lineTotal")
        or entry.get("LineTotal")
        or entry.get("SellingPriceTotal")
        or entry.get("Total")
    )
    if total_price is None:
        total_price = round(unit_price * quantity, 2)
    if unit_price <= 0 and quantity > 0 and total_price > 0:
        unit_price = round(total_price / quantity, 2)
    return {
        "name": name or default_kind.title(),
        "description": description,
        "quantity": max(quantity, 1),
        "unit_price": round(unit_price, 2),
        "total_price": round(total_price, 2),
    }
